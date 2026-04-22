"""
Integration tests for the Entomology API routes
(`/api/v1/entomology/snapshot`, `/issues`, `/report.xlsx`).

Uses the `collector` fixture from conftest plus a fresh EventLedger,
overrides auth + collector dependencies so tests don't need a real PIN.
"""

from __future__ import annotations

from io import BytesIO

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from app.api import dependencies as deps
from app.api.routes.auth import get_current_session
from app.api.routes import entomology as ent_routes
from app.models.diagnostic_event import DiagnosticCategory, DiagnosticSeverity


# ── Helpers ─────────────────────────────────────────────
def _fake_session():
    return {"employee_id": "E1", "name": "Tester", "roles": ["manager"]}


def _fake_probes():
    return [
        {
            "id": "database",
            "category": "CRITICAL",
            "status": "PASS",
            "message": None,
            "duration_ms": 5,
            "last_checked": "2026-04-21T00:00:00+00:00",
        }
    ]


# ── Fixtures ────────────────────────────────────────────
@pytest_asyncio.fixture
async def authed_client(collector, monkeypatch):
    """
    FastAPI client with:
      - DiagnosticCollector dependency → the conftest `collector` fixture
      - get_current_session dependency → fake session (no real auth)
      - `_run_probes_sync` stubbed to avoid loading kindnostic probes in tests
    """
    from app.main import app

    monkeypatch.setattr(ent_routes, "_run_probes_sync", _fake_probes)

    app.dependency_overrides[deps.get_diagnostic_collector] = lambda: collector
    app.dependency_overrides[get_current_session] = _fake_session

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac

    app.dependency_overrides.clear()


@pytest_asyncio.fixture
async def unauthed_client(collector, monkeypatch):
    """Same app but without overriding `get_current_session` — calls hit real auth."""
    from app.main import app

    monkeypatch.setattr(ent_routes, "_run_probes_sync", _fake_probes)
    app.dependency_overrides[deps.get_diagnostic_collector] = lambda: collector

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        yield ac

    app.dependency_overrides.clear()


# ── Tests ───────────────────────────────────────────────
@pytest.mark.asyncio
async def test_snapshot_requires_auth(unauthed_client):
    resp = await unauthed_client.get("/api/v1/entomology/snapshot")
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_issues_requires_auth(unauthed_client):
    resp = await unauthed_client.get("/api/v1/entomology/issues")
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_report_requires_auth(unauthed_client):
    resp = await unauthed_client.get("/api/v1/entomology/report.xlsx")
    assert resp.status_code == 401


@pytest.mark.asyncio
async def test_snapshot_shape(authed_client):
    resp = await authed_client.get("/api/v1/entomology/snapshot")
    assert resp.status_code == 200
    data = resp.json()
    assert data["terminal_id"] == "terminal-test-01"
    assert "generated_at" in data
    assert isinstance(data["probes"], list)
    assert data["probes"][0]["id"] == "database"
    assert "system_metrics" in data
    sm = data["system_metrics"]
    for key in ("memory_used_pct", "disk_used_pct", "cpu_temp_c", "uptime_hours"):
        assert key in sm


@pytest.mark.asyncio
async def test_issues_filters_by_severity_and_groups(authed_client, collector):
    # INFO event → should be filtered out
    await collector.record(
        category=DiagnosticCategory.SYSTEM,
        severity=DiagnosticSeverity.INFO,
        source="Test",
        event_code="SYS-HEARTBEAT",
        message="heartbeat",
        context={},
    )
    # WARNING event → included
    await collector.record(
        category=DiagnosticCategory.NETWORK,
        severity=DiagnosticSeverity.WARNING,
        source="Test",
        event_code="NET-007",
        message="latency",
        context={},
    )
    # CRITICAL event → included
    await collector.record(
        category=DiagnosticCategory.DEVICE,
        severity=DiagnosticSeverity.CRITICAL,
        source="Test",
        event_code="DEV-001",
        message="unreachable",
        context={},
    )

    resp = await authed_client.get("/api/v1/entomology/issues?days=1")
    assert resp.status_code == 200
    data = resp.json()

    assert data["total"] == 2
    groups = data["groups"]
    # All five categories present as keys
    for cat in ("DEVICE", "NETWORK", "SYSTEM", "PERIPHERAL", "RECOVERY"):
        assert cat in groups

    # DEVICE has DEV-001, NETWORK has NET-007
    assert "DEV-001" in groups["DEVICE"]
    assert "NET-007" in groups["NETWORK"]
    # SYSTEM should be empty (INFO excluded)
    assert groups["SYSTEM"] == {}


@pytest.mark.asyncio
async def test_report_xlsx_returns_valid_workbook(authed_client, collector):
    await collector.record(
        category=DiagnosticCategory.DEVICE,
        severity=DiagnosticSeverity.CRITICAL,
        source="Test",
        event_code="DEV-001",
        message="Terminal unreachable",
        context={"device_ip": "10.0.0.1"},
    )

    resp = await authed_client.get("/api/v1/entomology/report.xlsx?days=7")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp.headers["content-disposition"]
    assert ".xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(BytesIO(resp.content))
    assert "DEVICE Issues" in wb.sheetnames
    assert "Summary" in wb.sheetnames

    # Seeded event appears in DEVICE sheet messages (column E)
    device = wb["DEVICE Issues"]
    col_e = [device.cell(row=r, column=5).value for r in range(1, device.max_row + 1)]
    assert "Terminal unreachable" in col_e


@pytest.mark.asyncio
async def test_issues_rejects_invalid_days(authed_client):
    resp = await authed_client.get("/api/v1/entomology/issues?days=0")
    assert resp.status_code == 422
    resp = await authed_client.get("/api/v1/entomology/issues?days=9999")
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_issues_respects_min_severity(authed_client, collector):
    """`min_severity=INFO` must return events the WARNING+ default filters out.
    The entomology dashboard ships an "All (INFO+)" option so operators can
    see dead-end-tap traces (UI-007) and other INFO-level diagnostics."""
    # INFO event — dropped by the default WARNING+ filter, kept at INFO+
    await collector.record(
        category=DiagnosticCategory.UI,
        severity=DiagnosticSeverity.INFO,
        source="check-overview.handlePay",
        event_code="UI-007",
        message="dead-end tap",
        context={},
    )
    # WARNING event — always kept
    await collector.record(
        category=DiagnosticCategory.UI,
        severity=DiagnosticSeverity.WARNING,
        source="scene-manager.interrupt",
        event_code="UI-001",
        message="interrupt stacked",
        context={},
    )

    # Default — INFO event is dropped
    r1 = await authed_client.get("/api/v1/entomology/issues?days=1")
    assert r1.status_code == 200
    assert r1.json()["total"] == 1  # only UI-001

    # min_severity=INFO — both events kept
    r2 = await authed_client.get("/api/v1/entomology/issues?days=1&min_severity=INFO")
    assert r2.status_code == 200
    assert r2.json()["total"] == 2

    # Case-insensitive
    r3 = await authed_client.get("/api/v1/entomology/issues?days=1&min_severity=info")
    assert r3.status_code == 200
    assert r3.json()["total"] == 2

    # Invalid value falls back to WARNING (no 400)
    r4 = await authed_client.get("/api/v1/entomology/issues?days=1&min_severity=bogus")
    assert r4.status_code == 200
    assert r4.json()["total"] == 1


# ─── SEC-004 — forged-self terminal_id on sync replay ─────────────────

@pytest_asyncio.fixture
async def authed_client_with_ledger(authed_client, ledger, collector):
    """authed_client plus:
      - get_ledger override so sync/replay can resolve its EventLedger dep
      - DiagnosticCollector wired via set_diagnostic_collector so
        _record_diag (which pulls the module-level singleton, not a
        FastAPI Depends) can actually land SEC-004 into the test collector.
    """
    from app.main import app as _app
    _app.dependency_overrides[deps.get_ledger] = lambda: ledger
    prior = deps._diagnostic_collector  # type: ignore[attr-defined]
    deps.set_diagnostic_collector(collector)
    try:
        yield authed_client
    finally:
        deps._diagnostic_collector = prior  # type: ignore[attr-defined]


@pytest.mark.asyncio
async def test_sec004_emitted_when_replay_claims_this_terminal(
    authed_client_with_ledger, collector,
):
    """A replay batch containing events whose terminal_id equals this
    terminal's own id is suspicious — SEC-004 flags it."""
    from app.config import settings

    body = {"events": [
        {"event_type": "ROLE_CREATED", "terminal_id": settings.terminal_id,
         "event_id": "forged-1", "payload": {"role_id": "r1", "name": "X"}},
    ]}
    r = await authed_client_with_ledger.post("/api/v1/sync/config/events/replay", json=body)
    assert r.status_code == 200
    events = await collector.get_events(event_code="SEC-004")
    assert len(events) >= 1
    assert events[0].context["local_terminal_id"] == settings.terminal_id
    assert events[0].context["self_claim_count"] == 1


@pytest.mark.asyncio
async def test_sec004_quiet_when_replay_claims_only_foreign_ids(
    authed_client_with_ledger, collector,
):
    """Legitimate Overseer-origin replays (no self-claims) don't emit SEC-004."""
    body = {"events": [
        {"event_type": "ROLE_CREATED", "terminal_id": "OVERSEER",
         "event_id": "from-overseer-1",
         "payload": {"role_id": "r1", "name": "X"}},
    ]}
    r = await authed_client_with_ledger.post("/api/v1/sync/config/events/replay", json=body)
    assert r.status_code == 200
    assert len(await collector.get_events(event_code="SEC-004")) == 0
