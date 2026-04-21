"""
Unit tests for the Excel bug-report generator
(`app.services.entomology_report.build_bug_report_workbook`).

The generator is pure — no DB, no web layer — so these tests build
synthetic DiagnosticEvent objects directly and read the workbook back
with openpyxl.
"""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.diagnostic_event import (
    DiagnosticCategory,
    DiagnosticEvent,
    DiagnosticSeverity,
    GENESIS_HASH,
)
from app.services.entomology_report import (
    build_bug_report_workbook,
    workbook_to_bytes,
)


EXPECTED_SHEETS = [
    "Summary",
    "Current Snapshot",
    "DEVICE Issues",
    "NETWORK Issues",
    "SYSTEM Issues",
    "PERIPHERAL Issues",
    "RECOVERY Issues",
]


def _make_event(
    category: DiagnosticCategory,
    severity: DiagnosticSeverity,
    event_code: str,
    message: str = "msg",
    source: str = "TestSource",
    context: dict | None = None,
    timestamp: datetime | None = None,
) -> DiagnosticEvent:
    return DiagnosticEvent(
        terminal_id="terminal-test-01",
        timestamp=timestamp or datetime.now(timezone.utc),
        category=category,
        severity=severity,
        source=source,
        event_code=event_code,
        message=message,
        context=context or {"k": "v"},
        prev_hash=GENESIS_HASH,
        hash="a" * 64,
    )


def _basic_snapshot(generated_at: datetime) -> dict:
    return {
        "terminal_id": "terminal-test-01",
        "generated_at": generated_at,
        "heartbeat_age_seconds": 3.5,
        "probes": [
            {
                "id": "database",
                "status": "PASS",
                "message": None,
                "last_checked": generated_at,
            }
        ],
        "system_metrics": {
            "memory_used_pct": 42.0,
            "disk_used_pct": 22.0,
            "cpu_temp_c": 55.0,
            "uptime_hours": 6.0,
        },
    }


def _load(wb_bytes: bytes):
    return load_workbook(BytesIO(wb_bytes))


# ── Tests ─────────────────────────────────────────────────
def test_all_expected_sheets_present():
    now = datetime.now(timezone.utc)
    wb = build_bug_report_workbook(
        events=[],
        snapshot=_basic_snapshot(now),
        date_range=(now - timedelta(days=7), now),
    )
    assert wb.sheetnames == EXPECTED_SHEETS


def test_empty_categories_still_render_headers_and_placeholder():
    now = datetime.now(timezone.utc)
    wb = build_bug_report_workbook(
        events=[],
        snapshot=_basic_snapshot(now),
        date_range=(now - timedelta(days=7), now),
    )
    data = workbook_to_bytes(wb)
    reloaded = _load(data)

    for cat_name in [
        "DEVICE Issues",
        "NETWORK Issues",
        "SYSTEM Issues",
        "PERIPHERAL Issues",
        "RECOVERY Issues",
    ]:
        sheet = reloaded[cat_name]
        # Row 1 header
        assert sheet.cell(row=1, column=1).value == "timestamp (UTC)"
        # Row 2 placeholder
        assert sheet.cell(row=2, column=1).value.startswith("No ")


def test_device_sheet_contains_event_rows_grouped_by_code():
    now = datetime.now(timezone.utc)
    events = [
        _make_event(
            DiagnosticCategory.DEVICE,
            DiagnosticSeverity.CRITICAL,
            "DEV-001",
            message="Terminal unreachable",
        ),
        _make_event(
            DiagnosticCategory.DEVICE,
            DiagnosticSeverity.ERROR,
            "DEV-001",
            message="Second instance",
        ),
        _make_event(
            DiagnosticCategory.DEVICE,
            DiagnosticSeverity.WARNING,
            "DEV-002",
            message="Terminal timeout",
        ),
        # Different category — should not appear in DEVICE sheet
        _make_event(
            DiagnosticCategory.NETWORK,
            DiagnosticSeverity.WARNING,
            "NET-007",
            message="Elevated latency",
        ),
    ]
    wb = build_bug_report_workbook(
        events=events,
        snapshot=_basic_snapshot(now),
        date_range=(now - timedelta(days=7), now),
    )
    data = workbook_to_bytes(wb)
    reloaded = _load(data)
    sheet = reloaded["DEVICE Issues"]

    # Flatten column A values for easy search
    col_a = [sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1)]
    # Header row
    assert col_a[0] == "timestamp (UTC)"
    # Divider rows exist, one per event_code (DEV-001, DEV-002)
    divider_rows = [v for v in col_a if isinstance(v, str) and v.startswith("DEV-")]
    assert "DEV-001  (2)" in divider_rows
    assert "DEV-002  (1)" in divider_rows

    # Collect messages (column E)
    col_e = [sheet.cell(row=r, column=5).value for r in range(1, sheet.max_row + 1)]
    assert "Terminal unreachable" in col_e
    assert "Second instance" in col_e
    assert "Terminal timeout" in col_e
    # Network event did not leak into DEVICE sheet
    assert "Elevated latency" not in col_e


def test_summary_counts_matrix():
    now = datetime.now(timezone.utc)
    events = [
        _make_event(DiagnosticCategory.DEVICE, DiagnosticSeverity.CRITICAL, "DEV-001"),
        _make_event(DiagnosticCategory.DEVICE, DiagnosticSeverity.ERROR, "DEV-001"),
        _make_event(DiagnosticCategory.NETWORK, DiagnosticSeverity.WARNING, "NET-001"),
    ]
    wb = build_bug_report_workbook(
        events=events,
        snapshot=_basic_snapshot(now),
        date_range=(now - timedelta(days=7), now),
    )
    reloaded = _load(workbook_to_bytes(wb))
    summary = reloaded["Summary"]

    # Find the "Total issues" row: it's a (label, value) pair
    total_value = None
    for row in summary.iter_rows(values_only=True):
        if row and row[0] == "Total issues":
            total_value = row[1]
            break
    assert total_value == "3"


def test_snapshot_sheet_has_probes_and_metrics():
    now = datetime.now(timezone.utc)
    wb = build_bug_report_workbook(
        events=[],
        snapshot=_basic_snapshot(now),
        date_range=(now - timedelta(days=1), now),
    )
    reloaded = _load(workbook_to_bytes(wb))
    snap = reloaded["Current Snapshot"]

    # Probes header
    assert snap.cell(row=1, column=1).value == "Probes"
    assert snap.cell(row=3, column=1).value == "database"
    assert snap.cell(row=3, column=2).value == "PASS"

    # Metrics are further down — look for any cell == "System metrics"
    found_metrics = False
    for row in snap.iter_rows(values_only=True):
        if row and row[0] == "System metrics":
            found_metrics = True
            break
    assert found_metrics
