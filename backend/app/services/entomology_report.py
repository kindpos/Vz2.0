"""
Entomology bug-report workbook generator.

Builds an .xlsx workbook that organizes diagnostic issues by "scene"
(DiagnosticCategory) and "type" (event_code). Pure function — no web layer,
no DB access. Callers pass in the already-filtered events and a snapshot dict.
"""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.diagnostic_event import (
    DiagnosticCategory,
    DiagnosticEvent,
    DiagnosticSeverity,
)


# ── Styling constants ──────────────────────────────────────────────────────
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="3B3A39")
_DIVIDER_FONT = Font(bold=True, color="5B4500")
_DIVIDER_FILL = PatternFill("solid", fgColor="F5A623")  # KIND gold
_SEVERITY_FILL = {
    DiagnosticSeverity.WARNING: PatternFill("solid", fgColor="FFE69C"),
    DiagnosticSeverity.ERROR: PatternFill("solid", fgColor="F4B183"),
    DiagnosticSeverity.CRITICAL: PatternFill("solid", fgColor="E8472A"),
}

_CATEGORY_SHEETS: list[DiagnosticCategory] = [
    DiagnosticCategory.DEVICE,
    DiagnosticCategory.NETWORK,
    DiagnosticCategory.SYSTEM,
    DiagnosticCategory.PERIPHERAL,
    DiagnosticCategory.RECOVERY,
]

_ISSUE_HEADERS = [
    "timestamp (UTC)",
    "severity",
    "event_code",
    "source",
    "message",
    "context",
    "diagnostic_id",
    "correlation_id",
]

# severity weighting: highest first in output
_SEV_WEIGHT = {
    DiagnosticSeverity.CRITICAL: 0,
    DiagnosticSeverity.ERROR: 1,
    DiagnosticSeverity.WARNING: 2,
    DiagnosticSeverity.INFO: 3,
}


# ── Public entrypoint ──────────────────────────────────────────────────────
def build_bug_report_workbook(
    events: Iterable[DiagnosticEvent],
    snapshot: dict,
    date_range: tuple[datetime, datetime],
) -> Workbook:
    """
    Build an .xlsx workbook organizing `events` by category (scene) and
    event_code (type), plus a Summary sheet and a Current Snapshot sheet.

    Args:
        events: diagnostic events to include. Caller is responsible for
            pre-filtering to the desired severity / date range.
        snapshot: current-state dict, shape roughly:
            {
                "terminal_id": str,
                "generated_at": datetime,
                "heartbeat_age_seconds": float | None,
                "probes": [ { "id", "status", "message", "last_checked" }, ... ],
                "system_metrics": {
                    "memory_used_pct", "disk_used_pct",
                    "cpu_temp_c", "uptime_hours",
                },
            }
        date_range: (since, until) used in the Summary header.

    Returns:
        An openpyxl Workbook ready to save or stream.
    """
    events_list = list(events)

    wb = Workbook()
    # openpyxl creates a default sheet we'll rename for the summary
    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_summary(summary_ws, events_list, snapshot, date_range)

    _write_snapshot(wb.create_sheet("Current Snapshot"), snapshot)

    for cat in _CATEGORY_SHEETS:
        sheet = wb.create_sheet(f"{cat.value} Issues")
        cat_events = [e for e in events_list if e.category == cat]
        _write_category_sheet(sheet, cat, cat_events)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize a workbook to bytes for streaming responses."""
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Sheet writers ──────────────────────────────────────────────────────────
def _write_summary(
    ws,
    events: list[DiagnosticEvent],
    snapshot: dict,
    date_range: tuple[datetime, datetime],
) -> None:
    terminal_id = snapshot.get("terminal_id", "unknown")
    generated_at = snapshot.get("generated_at") or datetime.utcnow()
    if isinstance(generated_at, datetime):
        generated_at_str = generated_at.isoformat()
    else:
        generated_at_str = str(generated_at)
    heartbeat_age = snapshot.get("heartbeat_age_seconds")
    since, until = date_range

    rows: list[tuple[str, str]] = [
        ("Terminal ID", terminal_id),
        ("Generated (UTC)", generated_at_str),
        ("Window start (UTC)", since.isoformat()),
        ("Window end (UTC)", until.isoformat()),
        ("Total issues", str(len(events))),
        (
            "Heartbeat age (s)",
            "n/a" if heartbeat_age is None else f"{heartbeat_age:.1f}",
        ),
    ]

    for row_ix, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_ix, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_ix, column=2, value=value)

    # Counts matrix: categories × severities
    start_row = len(rows) + 2
    ws.cell(row=start_row, column=1, value="Counts by category × severity").font = Font(
        bold=True, size=12
    )

    header_row = start_row + 1
    severities = [
        DiagnosticSeverity.CRITICAL,
        DiagnosticSeverity.ERROR,
        DiagnosticSeverity.WARNING,
        DiagnosticSeverity.INFO,
    ]

    ws.cell(row=header_row, column=1, value="category")
    for col_ix, sev in enumerate(severities, start=2):
        ws.cell(row=header_row, column=col_ix, value=sev.value)
    ws.cell(row=header_row, column=len(severities) + 2, value="total")

    for row_ix, cat in enumerate(_CATEGORY_SHEETS, start=header_row + 1):
        ws.cell(row=row_ix, column=1, value=cat.value).font = Font(bold=True)
        row_total = 0
        for col_ix, sev in enumerate(severities, start=2):
            count = sum(
                1 for e in events if e.category == cat and e.severity == sev
            )
            ws.cell(row=row_ix, column=col_ix, value=count)
            row_total += count
        ws.cell(
            row=row_ix, column=len(severities) + 2, value=row_total
        ).font = Font(bold=True)

    # Style the header row
    for col_ix in range(1, len(severities) + 3):
        cell = ws.cell(row=header_row, column=col_ix)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    _autosize(ws, max_width=40)


def _write_snapshot(ws, snapshot: dict) -> None:
    # Probes
    ws.cell(row=1, column=1, value="Probes").font = Font(bold=True, size=12)
    probe_headers = ["id", "status", "message", "last_checked"]
    for col_ix, h in enumerate(probe_headers, start=1):
        cell = ws.cell(row=2, column=col_ix, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    probes = snapshot.get("probes") or []
    for row_ix, probe in enumerate(probes, start=3):
        ws.cell(row=row_ix, column=1, value=probe.get("id", ""))
        ws.cell(row=row_ix, column=2, value=probe.get("status", ""))
        ws.cell(row=row_ix, column=3, value=probe.get("message", "") or "")
        last_checked = probe.get("last_checked")
        if isinstance(last_checked, datetime):
            last_checked = last_checked.isoformat()
        ws.cell(row=row_ix, column=4, value=last_checked or "")

    # System metrics
    metrics = snapshot.get("system_metrics") or {}
    metrics_start = max(len(probes) + 4, 5)
    ws.cell(row=metrics_start, column=1, value="System metrics").font = Font(
        bold=True, size=12
    )
    metric_rows = [
        ("memory_used_pct", metrics.get("memory_used_pct")),
        ("disk_used_pct", metrics.get("disk_used_pct")),
        ("cpu_temp_c", metrics.get("cpu_temp_c")),
        ("uptime_hours", metrics.get("uptime_hours")),
    ]
    for i, (label, value) in enumerate(metric_rows, start=1):
        ws.cell(row=metrics_start + i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=metrics_start + i, column=2, value=value)

    _autosize(ws, max_width=60)


def _write_category_sheet(
    ws,
    category: DiagnosticCategory,
    events: list[DiagnosticEvent],
) -> None:
    # Header row
    for col_ix, h in enumerate(_ISSUE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_ix, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    if not events:
        ws.cell(
            row=2,
            column=1,
            value=f"No {category.value} issues in this window.",
        ).font = Font(italic=True, color="888888")
        _autosize(ws, max_width=80)
        return

    # Group by event_code, sort codes alphabetically
    grouped: dict[str, list[DiagnosticEvent]] = {}
    for e in events:
        grouped.setdefault(e.event_code, []).append(e)

    for rows in grouped.values():
        rows.sort(
            key=lambda e: (_SEV_WEIGHT.get(e.severity, 99), -e.timestamp.timestamp())
        )

    current_row = 2
    for code in sorted(grouped.keys()):
        # Divider row for this event_code group
        divider_cell = ws.cell(
            row=current_row,
            column=1,
            value=f"{code}  ({len(grouped[code])})",
        )
        divider_cell.font = _DIVIDER_FONT
        divider_cell.fill = _DIVIDER_FILL
        for col_ix in range(2, len(_ISSUE_HEADERS) + 1):
            ws.cell(row=current_row, column=col_ix).fill = _DIVIDER_FILL
        current_row += 1

        for ev in grouped[code]:
            ws.cell(row=current_row, column=1, value=ev.timestamp.isoformat())
            sev_cell = ws.cell(row=current_row, column=2, value=ev.severity.value)
            fill = _SEVERITY_FILL.get(ev.severity)
            if fill is not None:
                sev_cell.fill = fill
            ws.cell(row=current_row, column=3, value=ev.event_code)
            ws.cell(row=current_row, column=4, value=ev.source)
            ws.cell(row=current_row, column=5, value=ev.message)
            ws.cell(
                row=current_row,
                column=6,
                value=json.dumps(ev.context, sort_keys=True, default=str),
            )
            ws.cell(row=current_row, column=7, value=ev.diagnostic_id)
            ws.cell(row=current_row, column=8, value=ev.correlation_id or "")
            current_row += 1

    _autosize(ws, max_width=80)


# ── Helpers ────────────────────────────────────────────────────────────────
def _autosize(ws, max_width: int = 60) -> None:
    """Rudimentary column autosize — openpyxl has no built-in."""
    for col_cells in ws.columns:
        longest = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            length = len(str(value))
            if length > longest:
                longest = length
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)
